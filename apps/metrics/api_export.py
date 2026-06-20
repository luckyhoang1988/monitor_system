"""API xuất dữ liệu Metrics ra CSV và Excel."""
import csv
from django.http import HttpResponse
from django.utils import timezone
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from apps.devices.models import Device, Interface
from apps.metrics.models import (
    SystemHealth, SystemHealthHourly, SystemHealthDaily,
    InterfaceStats, InterfaceStatsHourly, InterfaceStatsDaily
)
from apps.metrics.api import _parse_range
import openpyxl
from openpyxl.utils import get_column_letter

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def export_metrics(request):
    """
    Export metrics (System Health hoặc Interface) ra CSV/Excel.
    Query params:
    - device_id: ID của device
    - type: 'system' hoặc 'interface'
    - range: '1h', '24h', '7d', '30d'
    - export_format: 'csv' hoặc 'excel' (default: csv)
    """
    device_id = request.GET.get("device_id")
    export_type = request.GET.get("type", "system")
    range_str = request.GET.get("range", "24h")
    export_format = request.GET.get("export_format", "csv")

    if not device_id:
        return HttpResponse("Thiếu device_id", status=400)

    try:
        device = Device.objects.get(pk=device_id)
    except Device.DoesNotExist:
        return HttpResponse("Device không tồn tại", status=404)

    delta, source = _parse_range(range_str)
    since = timezone.now() - delta

    if export_type == "system":
        headers, rows = _get_system_export_data(device, source, since)
        filename = f"System_Health_{device.name}_{range_str}"
    elif export_type == "interface":
        headers, rows = _get_interface_export_data(device, source, since)
        filename = f"Interfaces_{device.name}_{range_str}"
    else:
        return HttpResponse("Type không hợp lệ", status=400)

    if export_format == "excel":
        return _export_excel(headers, rows, filename)
    else:
        return _export_csv(headers, rows, filename)


def _get_system_export_data(device, source, since):
    headers = ["Thời gian", "CPU (%)", "Memory (%)"]
    rows = []

    if source == "daily":
        qs = SystemHealthDaily.objects.filter(device=device, day__gte=since.date()).order_by("day")
        headers.extend(["CPU Max", "Memory Max"])
        for r in qs:
            rows.append([r.day.strftime("%Y-%m-%d"), r.cpu_avg, r.mem_avg, r.cpu_max, r.mem_max])
    elif source == "hourly":
        qs = SystemHealthHourly.objects.filter(device=device, hour__gte=since).order_by("hour")
        headers.extend(["CPU Max", "Memory Max"])
        for r in qs:
            rows.append([r.hour.strftime("%Y-%m-%d %H:00"), r.cpu_avg, r.mem_avg, r.cpu_max, r.mem_max])
    else:
        qs = SystemHealth.objects.filter(device=device, timestamp__gte=since).order_by("timestamp")
        for r in qs:
            rows.append([r.timestamp.strftime("%Y-%m-%d %H:%M:%S"), r.cpu_percent, r.mem_percent])
            
    return headers, rows


def _get_interface_export_data(device, source, since):
    headers = ["Interface", "Thời gian", "In Mbps", "Out Mbps"]
    rows = []
    interfaces = Interface.objects.filter(device=device)

    if source == "daily":
        headers.extend(["In Max", "Out Max", "In Errors", "Out Errors"])
        for iface in interfaces:
            qs = InterfaceStatsDaily.objects.filter(interface=iface, day__gte=since.date()).order_by("day")
            for r in qs:
                rows.append([iface.name, r.day.strftime("%Y-%m-%d"), r.in_mbps_avg, r.out_mbps_avg, r.in_mbps_max, r.out_mbps_max, r.in_errors, r.out_errors])
    elif source == "hourly":
        headers.extend(["In Max", "Out Max", "In Errors", "Out Errors"])
        for iface in interfaces:
            qs = InterfaceStatsHourly.objects.filter(interface=iface, hour__gte=since).order_by("hour")
            for r in qs:
                rows.append([iface.name, r.hour.strftime("%Y-%m-%d %H:00"), r.in_mbps_avg, r.out_mbps_avg, r.in_mbps_max, r.out_mbps_max, r.in_errors, r.out_errors])
    else:
        headers.extend(["In Errors", "Out Errors"])
        for iface in interfaces:
            qs = InterfaceStats.objects.filter(interface=iface, timestamp__gte=since).order_by("timestamp")
            for r in qs:
                rows.append([iface.name, r.timestamp.strftime("%Y-%m-%d %H:%M:%S"), r.in_mbps, r.out_mbps, r.in_errors, r.out_errors])

    return headers, rows


def _export_csv(headers, rows, filename):
    response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'

    writer = csv.writer(response)
    writer.writerow(headers)
    for row in rows:
        writer.writerow(row)
    return response


def _export_excel(headers, rows, filename):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Metrics Data"

    ws.append(headers)
    for row in rows:
        ws.append(row)

    # Auto-adjust column width
    for i in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 15

    wb.save(response)
    return response
